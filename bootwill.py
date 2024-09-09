from tkinter import *
from tkinter import ttk

from openpyxl import load_workbook

from time import strftime
from translate import Translator

import functions

root = Tk()

root.geometry('700x500')
root.title('Bootwill')
root.resizable(False, False)

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
            'T', 'U', 'V', 'W', 'X', 'Y']

dict_info = {
    'month': '',
    'place': '',
    'date': '',
    'invoice': '',
    'plan_date': '',
    'fakt_date': '',
    'customer': '',
    'phone': '',
    'subject': '',
    'color': '',
    'material': '',
    'work': '',
    'product': '',
    'notes': '',
    'couple': '',
    'cost': '',
    'payment': '',
    'itog_skid': '',
    'calculation': '',
    'using': '',
    'calc_with_masters': '',
    'FOT': '',
    'delivery': '',
    'add_cost': '',
    'sum_cost': '',
    'income_expense': ''
}

translator = Translator(from_lang='en', to_lang='ru')
current_month = translator.translate(strftime('%B'))


def import_information(dict_info):
    wb = load_workbook('info.xls')

    names = wb.sheetnames

    if month_entry.get() not in names:
        sheet = wb.create_sheet(str(month_entry.get()))
    else:
        sheet = wb[str(month_entry.get())]

    for i in range(4, 10000):
        if sheet.cell(row=i, column=2).value == None:
            latest_num = i
            break

    for i in range(0, len(alphabet)):
        coordinate = alphabet[i] + str(latest_num)
        sheet[coordinate] = dict_info[dict_info.keys()[i+1]]

    wb.save('info.xls')


def updating_information():
    month_entry.delete(0, 'end')
    month_entry.insert(0, current_month)
    place_combo.delete('0', 'end')
    place_combo.current(0)
    date_of_admission_entry.delete(0, 'end')
    invoice_entry.delete(0, 'end')
    date_issue_plan_entry.delete(0, 'end')
    date_issue_fakt_entry.delete(0, 'end')
    customer_entry.delete(0, 'end')
    phone_entry.delete(0, 'end')
    subject_repair_entry.delete(0, 'end')
    color_entry.delete(0, 'end')
    material_combo.delete('0', 'end')
    material_combo.current(0)
    product_combo.delete('0', 'end')
    product_combo.current(0)
    order_notes_text.delete("1.0", "end")
    count_couple_combo.delete('0', 'end')
    count_couple_combo.current(0)
    cost_entry.delete(0, 'end')
    payment_combo.delete('0', 'end')
    payment_combo.current(0)
    calculation_combo.delete('0', 'end')
    calculation_combo.current(1)
    use_combo.delete('0', 'end')
    use_combo.current(0)
    delivery_entry.delete(0, 'end')
    add_costs_combo.delete('0', 'end')
    add_costs_combo.current(0)


def saving_changes():
    global dict_info
    dict_info['month'] = month = month_entry.get()
    dict_info['place'] = place = place_combo.get()
    dict_info['date'] = date = date_of_admission_entry.get()
    dict_info['invoice'] = invoice = invoice_entry.get()
    dict_info['plan_date'] = plan_date = date_issue_plan_entry.get()
    dict_info['fakt_date'] = fakt_date = date_issue_fakt_entry.get()
    dict_info['customer'] = customer = customer_entry.get()
    dict_info['phone'] = phone = phone_entry.get()
    dict_info['subject'] = subject = subject_repair_entry.get()
    dict_info['color'] = color = color_entry.get()
    dict_info['material'] = material = material_combo.get()
    if r_var.get() == 0:
        dict_info['work'] = work = 'Ремонт'
    elif r_var.get() == 1:
        dict_info['work'] = work = 'Чистка'
    else:
        dict_info['work'] = work = '---'
    dict_info['product'] = product = product_combo.get()
    if len(order_notes_text.get(index1=0.0, index2=15.0)[:-2:]) == 0:
        dict_info['notes'] = notes = '---'
    else:
        dict_info['notes'] = notes = order_notes_text.get(index1=0.0, index2=15.0)[:-1:]
    dict_info['couple'] = couple = count_couple_combo.get()
    dict_info['cost'] = cost = cost_entry.get()
    dict_info['payment'] = payment = payment_combo.get()
    dict_info['itog_skid'] = itog_skid = str(int(couple) * int(cost))
    dict_info['calculation'] = calculation = calculation_combo.get()
    dict_info['using'] = using = use_combo.get()
    if work == 'Ремонт':
        dict_info['calc_with_masters'] = calc_with_masters = str(int(itog_skid) // 2)
    else:
        dict_info['calc_with_masters'] = calc_with_masters = 0

    if work == 'Чистка':
        dict_info['FOT'] = FOT = str(int(itog_skid) // 100 * 40)
    else:
        dict_info['FOT'] = FOT = str(int(itog_skid) // 100 * 10)
    dict_info['delivery'] = delivery = int(delivery_entry.get())
    dict_info['add_cost'] = add_cost = int(add_costs_combo.get())
    dict_info['sum_cost'] = sum_cost = str(int(delivery) + int(add_cost))

    if calculation == '100':
        dict_info['income_expense'] = str((int(itog_skid) // 1) - int(FOT) - int(delivery))
    else:
        dict_info['income_expense'] = str((int(itog_skid) // 2) - int(FOT) - int(delivery))

    if TRUE:
        after_save_btn['fg'] = 'Green'
        after_save_btn['text'] = 'Изменения сохранены успешно'
    else:
        after_save_btn['fg'] = 'Red'
        after_save_btn['text'] = 'Ошибка с введёнными данными'

    import_information(dict_info)

    updating_information()


# save info
save_btn = Button(root, text='Save changes', command=saving_changes)
save_btn.configure(width=20, height=2, bg='grey', fg='white')
save_btn.place(x=530, y=450)

# check saving
after_save_btn = Label(root, text='')
after_save_btn.configure(font=19)
after_save_btn.place(x=445, y=400)

#input indormation
month_lbl = Label(root, text='Месяц: ', font=19)
month_lbl.place(x=20, y=25)

month_entry = Entry(root, width = 23)
month_entry.insert(0, current_month)
month_entry.place(x=125, y=26)


place_lbl = Label(root, text='Откуда заказ: ', font=19)
place_lbl.place(x=20, y=55)

place_combo = ttk.Combobox(root, values=["Сити"])
place_combo.current(0)
place_combo.place(x=125, y=56)


date_of_admission_lbl = Label(root, text='Дата приёма: ', font=19)
date_of_admission_lbl.place(x=20, y=85)

date_of_admission_entry = Entry(root, width=23)
date_of_admission_entry.place(x=125, y=86)


invoice_lbl = Label(root, text='Накладная: ', font=19)
invoice_lbl.place(x=20, y=115)

invoice_entry = Entry(root, width=23)
invoice_entry.place(x=125, y=116)


date_issue_plan_lbl = Label(root, text='Дата выдачи(план): ', font=19)
date_issue_plan_lbl.place(x=20, y=145)

date_issue_plan_entry = Entry(root, width=16)
date_issue_plan_entry.place(x=167, y=146)


date_issue_fakt_lbl = Label(root, text='Дата выдачи(факт): ', font=19)
date_issue_fakt_lbl.place(x=20, y=175)

date_issue_fakt_entry = Entry(root, width=16)
date_issue_fakt_entry.place(x=168, y=176)


customer_lbl = Label(root, text='Заказчик: ', font=19)
customer_lbl.place(x=20, y=205)

customer_entry = Entry(root, width=23)
customer_entry.place(x=125, y=206)


phone_lbl = Label(root, text='Телефон: ', font=19)
phone_lbl.place(x=20, y=235)

phone_entry = Entry(root, width=23)
phone_entry.place(x=125, y=236)


subject_repair_lbl = Label(root, text='Предмет ремонта: ', font=19)
subject_repair_lbl.place(x=20, y=265)

subject_repair_entry = Entry(root, width=16)
subject_repair_entry.place(x=168, y=266)


color_lbl = Label(root, text='Цвет: ', font=19)
color_lbl.place(x=20, y=295)

color_entry = Entry(root, width=23)
color_entry.place(x=125, y=296)


material_lbl = Label(root, text='Материал: ', font=19)
material_lbl.place(x=20, y=325)

material_combo = ttk.Combobox(root, values=['нат. кожа', 'лак. кожа', 'замша', 'текстиль', 'синтетика',
                                            'ткань джинс', 'комби', 'шлиф кожа', 'ткань', 'нат. нубук'])
material_combo.current(0)
material_combo.place(x=125, y=326)


work_lbl = Label(root, text='Работы: ', font=19)
work_lbl.place(x=20, y=356)

r_var = IntVar()
r_var.set(0)
work_repair_radio = Radiobutton(text='Ремонт', font=15, variable=r_var, value=0)
work_repair_radio.place(x=125, y=356)
work_clean_radio = Radiobutton(text='Чистка', font=15, variable=r_var, value=1)
work_clean_radio.place(x=125, y=376)
work_none_radio = Radiobutton(text='---', font=15, variable=r_var, value=2)
work_none_radio.place(x=125, y=396)


product_lbl = Label(root, text='Товары: ', font=19)
product_lbl.place(x=20, y=425)

product_combo = ttk.Combobox(root, values=['---'])
product_combo.current(0)
product_combo.place(x=125, y=426)


order_notes_lbl = Label(root, text='Примечания к заказу: ', font=19)
order_notes_lbl.place(x=350, y=35)

order_notes_text = Text(root, width=40, height=7, wrap=WORD)
order_notes_text.place(x=350, y=56)


count_couple_lbl = Label(root, text='Кол-во пар: ', font=19)
count_couple_lbl.place(x=350, y=176)

count_couple_combo = ttk.Combobox(root, values=['1'])
count_couple_combo.current(0)
count_couple_combo.place(x=452, y=176)


cost_lbl = Label(root, text='Стоимость: ', font=19)
cost_lbl.place(x=350, y=206)

cost_entry = Entry(root, width=23)
cost_entry.place(x=452, y=207)


payment_lbl = Label(root, text='Форма расчёта: ', font=19)
payment_lbl.place(x=350, y=236)

payment_combo = ttk.Combobox(root, width=16, values=['---', 'Н', 'Б/Н'])
payment_combo.current(0)
payment_combo.place(x=472, y=237)


calculation_lbl = Label(root, text='Расчёт %: ', font=19)
calculation_lbl.place(x=350, y=266)

calculation_combo = ttk.Combobox(root, values=['100', '50/50'])
calculation_combo.current(1)
calculation_combo.place(x=452, y=267)


use_lbl = Label(root, text='Исп-ль: ', font=19)
use_lbl.place(x=350, y=296)

use_combo = ttk.Combobox(root, values=['БУТ', 'СМ', 'М/С'])
use_combo.current(0)
use_combo.place(x=452, y=297)


delivery_lbl = Label(root, text='Доставка: ', font=19)

delivery_lbl.place(x=350, y=326)

delivery_entry = Entry(root, width=23)
delivery_entry.insert(0, '0')
delivery_entry.place(x=452, y=327)


add_costs_lbl = Label(root, text='Доп.затраты: ', font=19)
add_costs_lbl.place(x=350, y=356)

add_costs_combo = ttk.Combobox(root, values=['0'])
add_costs_combo.current(0)
add_costs_combo.place(x=452, y=357)

canvas = Canvas(root, width=25, height=475)
canvas.place(x=300, y=12)
canvas.create_line(10, 25, 10, 450)

root.mainloop()
